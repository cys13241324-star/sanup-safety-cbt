# -*- coding: utf-8 -*-
"""산출물 점검 — 만들어 놓고 믿지 않는다."""
import collections
import glob
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import dataset
import build_all

OUT = r'D:\safety-cbt'


def main():
    recs = dataset.build()
    by = collections.defaultdict(list)
    for r in recs:
        by[r['round']].append(r)
    bad, note_only = [], []

    # 1) 정답·발문이 원고(ps1)와 재배열 메타(json)에서 서로 맞는가
    meta = dataset.load_meta()
    mism_a = mism_t = nometa = 0
    for r in recs:
        m = meta.get((r['round'], r['n']))
        if not m:
            nometa += 1
            continue
        if m['a'] != r['a']:
            mism_a += 1
            bad.append('정답 불일치 %s  원고%d vs 메타%d' % (r['code'], r['a'], m['a']))
        if re.sub(r'\s', '', m['t'])[:30] != re.sub(r'\s', '', r['t'])[:30]:
            mism_t += 1
            note_only.append(r['code'])
    print('두 소스 대조 — 메타없음 %d · 정답불일치 %d · 발문 글자차 %d(저자 정리분, 결함 아님)'
          % (nometa, mism_a, mism_t))

    # 2) 회차 폴더 구성
    for rd in sorted(by):
        y, rn = dataset.round_label(rd)
        d = os.path.join(OUT, dataset.folder(rd))
        xl = os.path.join(d, '산업안전기사_%d_%d회_문항등록.xlsx' % (y, rn))
        ht = os.path.join(d, '%d_%d회_검수본.html' % (y, rn))
        for p in (xl, ht, os.path.join(d, 'img'), os.path.join(d, 'katex')):
            if not os.path.exists(p):
                bad.append('없음 %s' % p)
        want = sum(1 for r in by[rd] if r['fig']) + sum(1 for r in by[rd] if r['qfig'])
        have = len(glob.glob(os.path.join(d, 'img', '*.png')))
        if want != have:
            bad.append('그림 개수 %s  기대%d 실제%d' % (dataset.folder(rd), want, have))
        html = open(ht, encoding='utf-8').read()
        nq = html.count('<div class=q>')
        if nq != 120:
            bad.append('검수본 문항수 %s = %d' % (dataset.folder(rd), nq))
        if '#=' in html or '}#' in html or '==' in html or '__' in html:
            bad.append('원고 표기 잔재 %s' % dataset.folder(rd))

    # 3) 엑셀 — 통합본 전수
    wb = openpyxl.load_workbook(os.path.join(OUT, '산업안전기사_CBT_24회차_통합.xlsx'),
                                read_only=True)
    ws = wb['문항등록']
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    rows = [r for r in rows if r[0]]
    print('통합 엑셀 %d행 × %d열' % (len(rows), ws.max_column))
    if len(rows) != len(recs):
        bad.append('통합 행수 %d != %d' % (len(rows), len(recs)))
    idx = {v: i for i, v in enumerate(build_all.HEAD)}
    req = ['과정', '연도', '회차', '번호', '문항코드', '과목ID', '발문',
           '보기1', '보기2', '보기3', '보기4', '정답(1~4)', '해설']
    empty = collections.Counter()
    codes = set()
    for row in rows:
        for k in req:
            if row[idx[k]] in (None, ''):
                empty[k] += 1
        codes.add(row[idx['문항코드']])
        if row[idx['정답(1~4)']] not in (1, 2, 3, 4):
            bad.append('정답값 이상 %s' % row[idx['문항코드']])
    if empty:
        bad.append('필수열 빈칸 %s' % dict(empty))
    if len(codes) != len(rows):
        bad.append('문항코드 중복 %d' % (len(rows) - len(codes)))
    print('문항코드 유일 %d' % len(codes))

    # 4) 선택열 채움률
    wb2 = openpyxl.load_workbook(os.path.join(OUT, '산업안전기사_CBT_24회차_통합.xlsx'),
                                 read_only=True)
    ws2 = wb2['문항등록']
    fill = collections.Counter()
    n = 0
    for row in ws2.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        n += 1
        for k in ('챕터', '대유형', '중유형', '내용', '오답분석', '학습포인트',
                  '조건', '발문그림', '비고'):
            if row[idx[k]] not in (None, ''):
                fill[k] += 1
    print('채움률 (%d행 기준)' % n)
    for k, v in fill.most_common():
        print('   %-8s %5d  %5.1f%%' % (k, v, 100 * v / n))

    # 5) CBT 화면 · 암기카드 · 링크
    import cards as cardmod
    cardlist, bycode = cardmod.build(recs)
    ids = {c['id'] for c in cardlist}
    print('암기카드 %d장 · 문항 연결 %d / %d' % (len(cardlist), len(bycode), len(recs)))
    if len(bycode) != len(recs):
        bad.append('카드 미연결 문항 %d' % (len(recs) - len(bycode)))
    if any(not c['front'] or not c['back'] for c in cardlist):
        bad.append('앞·뒷면 빈 카드 있음')

    for p in ('_암기카드.html', '_index.html', 'katex',
              '산업안전기사_암기카드_%d장.xlsx' % len(cardlist)):
        if not os.path.exists(os.path.join(OUT, p)):
            bad.append('없음 %s' % p)

    linked = 0
    for rd in sorted(by):
        y, rn = dataset.round_label(rd)
        p = os.path.join(OUT, dataset.folder(rd), '%d_%d회_CBT.html' % (y, rn))
        st = os.path.join(OUT, dataset.folder(rd), '%d_%d회_학습.html' % (y, rn))
        for q in (p, st):
            if not os.path.exists(q):
                bad.append('없음 %s' % q)
        if not os.path.exists(p) or not os.path.exists(st):
            continue
        html = open(p, encoding='utf-8').read()
        sh = open(st, encoding='utf-8').read()
        if 'const MODE="exam"' not in html:
            bad.append('시험 화면 모드 표시 없음 %s' % dataset.folder(rd))
        if 'const MODE="study"' not in sh:
            bad.append('학습 화면 모드 표시 없음 %s' % dataset.folder(rd))
        for m in re.finditer(r'"card": ?"(card_[a-z]+_t\d+)"', html):
            linked += 1
            if m.group(1) not in ids:
                bad.append('깨진 카드링크 %s' % m.group(1))
        rv = os.path.join(OUT, dataset.folder(rd), '%d_%d회_검수본.html' % (y, rn))
        for m in re.finditer(r'_암기카드\.html#(card_[a-z]+_t\d+)', open(rv, encoding='utf-8').read()):
            linked += 1
            if m.group(1) not in ids:
                bad.append('깨진 카드링크(검수본) %s' % m.group(1))
        if html.count('"n":') < 120:
            bad.append('CBT 문항수 부족 %s' % dataset.folder(rd))
    print('CBT 화면 24개 + 학습 화면 24개 · 카드 링크 %d개' % linked)

    # 6) 법령 원문
    import lawcite
    nref = nart = 0
    dead = 0
    for r in recs:
        rs = lawcite.refs(*lawcite.cite_source(r))
        rs = [x for x in rs if x[0] in lawcite.laws()]
        if not rs:
            continue
        nref += 1
        if any(k == 'jo' and lawcite.article(s0, n0, b0) for s0, k, n0, b0 in rs):
            nart += 1
        blk = lawcite.block(r)
        if blk and ('<details' not in blk or '</details>' not in blk):
            dead += 1
    print('법령 인용 문항 %d · 조문 원문이 붙은 문항 %d' % (nref, nart))
    if dead:
        bad.append('법령 블록 깨짐 %d' % dead)
    eff = {k: lawcite.effdate(k) for k in lawcite.laws()}
    if any(not v for v in eff.values()):
        bad.append('시행일 없는 법령 %s' % [k for k, v in eff.items() if not v])
    print('법령 %d종 시행일 — %s' % (
        len(eff), ' · '.join('%s %s' % (k, v) for k, v in list(eff.items())[:4])))

    # 엑셀의 관련카드ID
    filled = sum(1 for row in rows if row[idx['관련카드ID']])
    print('관련카드ID 채움 %d / %d' % (filled, len(rows)))
    if filled != len(rows):
        bad.append('관련카드ID 빈칸 %d' % (len(rows) - filled))

    print()
    if bad:
        print('문제 %d건' % len(bad))
        for b in bad[:40]:
            print('  -', b)
    else:
        print('점검 통과 — 문제 없음')


if __name__ == '__main__':
    main()
