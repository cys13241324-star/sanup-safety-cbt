# -*- coding: utf-8 -*-
"""산업안전기사 기출 재배열 24회 -> 전기기능사 CBT 포맷 산출물 전체.

만드는 것
  CBT_2026_1회 ~ CBT_2021_4회 /
      산업안전기사_{연도}_{회}회_문항등록.xlsx   v3 35열 등록양식
      {연도}_{회}회_검수본.html                  KaTeX 렌더 검수본
      img/                                       발문·해설 그림
      katex/                                     오프라인 수식 자원
  산업안전기사_CBT_24회차_통합.xlsx
  _index_검수본.html
  README.md
"""
import os
import re
import shutil
import sys
import collections

# 한글 콘솔(cp949)에서 em대시 같은 글자에 걸려 마지막 줄이 터지는 것을 막는다.
try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import dataset
import render
import cards as cardmod
import cbt_app
import card_view
import lawcite
from render import rich, table, wrongs, point, esc, plain

# 내보낼 자리와 물려받을 서식 — 컴퓨터마다 자리가 달라 있는 쪽을 고른다.
HERE = os.path.dirname(os.path.abspath(__file__))
OUT = dataset.pick(r'D:\safety-cbt', os.path.dirname(HERE))
KATEX_SRC = dataset.pick(r'D:\electrician-cbt\CBT_2025_1회\katex',
                         os.path.join(OUT, 'katex'))
# 등록양식은 전기기능사 엑셀에서 물려받았다.  그 파일이 없으면 이미 만들어 둔
# 산업안전기사 회차 엑셀에서 물려받는다 — 서식이 같으므로 결과가 같다.
DONOR = dataset.pick(
    r'D:\electrician-cbt\CBT_2025_1회\전기기능사_2025_1회_문항등록.xlsx',
    os.path.join(OUT, 'CBT_2026_1회', '산업안전기사_2026_1회_문항등록.xlsx'))
TEMPLATE = os.path.join(OUT, 'build', '_template.xlsx')

GROUPS = [('출처', 1, 6, '475569'), ('코드', 7, 7, '4F46E5'), ('강의', 8, 8, '7C3AED'),
          ('암기카드', 9, 9, 'C026D3'), ('시뮬레이터', 10, 10, '0284C7'),
          ('분류', 11, 15, '0891B2'), ('속성', 16, 20, 'E11D48'),
          ('발문', 21, 23, 'DB2777'), ('보기', 24, 31, '2563EB'),
          ('정답', 32, 32, '0D9488'), ('해설', 33, 33, 'EA580C'),
          ('오답분석', 34, 34, '9333EA'), ('학습POINT', 35, 35, '059669')]
HEAD = ['과정', '연도', '회차', '번호', '사용교재', '교재구분', '문항코드', '강의주소',
        '관련카드ID', '관련시뮬레이터ID', '과목ID', '챕터', '대유형', '중유형', '내용',
        '빈출도', '난이도', '문제유형', '변형이력', '비고', '발문', '조건', '발문그림',
        '보기1', '보기1그림', '보기2', '보기2그림', '보기3', '보기3그림', '보기4',
        '보기4그림', '정답(1~4)', '해설', '오답분석', '학습포인트']
SUBFILL = ['F1F5F9'] * 6 + ['E0E7FF', 'EDE9FE', 'FAE8FF', 'E0F2FE'] + \
          ['CFFAFE'] * 5 + ['FFE4E6'] * 5 + ['FCE7F3', 'FCE7F3', 'FEF3C7'] + \
          ['DBEAFE', 'FEF3C7'] * 4 + ['CCFBF1', 'FFEDD5', 'F3E8FF', 'D1FAE5']
WIDTH = [11, 7, 6, 6, 20, 8, 22, 28, 26, 28, 11, 14, 18, 18, 24, 8, 8, 10, 8, 18,
         45, 22, 16, 24, 13, 24, 13, 24, 13, 24, 13, 8, 55, 45, 55]

FLAGLABEL = [('revised', '원본 수정'), ('needfig', '그림 보강 필요'),
             ('suspect', '점검 요망'), ('crossed', '교차 오염 의심')]


# ------------------------------------------------------------------ 엑셀

def make_template():
    """전기기능사 등록양식을 서식째 물려받아 빈 틀을 만든다."""
    wb = openpyxl.load_workbook(DONOR)
    ws = wb['문항등록']
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)
    cur = wb['커리큘럼ID']
    cur.delete_rows(2, max(cur.max_row - 1, 1))
    for i, s in enumerate(dataset.SUBJ):
        cur.cell(row=i + 2, column=1, value='제%d과목' % (i + 1))
        cur.cell(row=i + 2, column=2, value=s)
    cur.cell(row=1, column=3, value='챕터(재배열 메타 기준)')
    wb.save(TEMPLATE)


def header(ws):
    for name, c1, c2, rgb in GROUPS:
        cell = ws.cell(row=1, column=c1, value=name)
        cell.fill = PatternFill('solid', fgColor=rgb)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if c2 > c1:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
    for i, name in enumerate(HEAD):
        cell = ws.cell(row=2, column=i + 1, value=name)
        cell.fill = PatternFill('solid', fgColor=SUBFILL[i])
        cell.font = Font(bold=True, color='1F2937')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i + 1)].width = WIDTH[i]
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = 'H3'


def _one_flag(v):
    """재배열 메타의 표시값은 참·거짓, 글줄, 개정내역 어느 쪽이든 올 수 있다."""
    if v is True or v == 1:
        return ''
    if isinstance(v, dict):
        what, src = v.get('what', ''), v.get('src', '')
        return '%s (%s)' % (what, src) if src else what
    if isinstance(v, list):
        return ' / '.join(x for x in (_one_flag(i) for i in v) if x)
    return str(v)


def flagtext(r):
    out = []
    for k, label in FLAGLABEL:
        v = r['flags'].get(k)
        if not v:
            continue
        txt = _one_flag(v)
        out.append('%s: %s' % (label, txt) if txt else label)
    return out


def note(r):
    bits = ['원본 시행일 %s (%d회 출제)' % (r['src'], r['rep']),
            '자동판정 빈출도%d · %s · 난이도%d' % (r['freq'], r['kind'], r['level'])]
    bits += flagtext(r)
    if not r['chapter']:
        bits.append('분류 미상')
    return ' | '.join(bits)


def solution_html(r):
    buf = [rich(r['sol'])]
    if r['tbl']:
        buf.append(table(r['tbl']))
    if r['fig']:
        buf.append('<img src="%s_exp1.png">' % r['code'])
        if r['figcap']:
            buf.append('<em>%s</em>' % rich(r['figcap'], False))
    return ''.join(buf)


def row_values(r):
    return [
        dataset.COURSE, r['year'], r['rno'], r['n'], dataset.BOOK, dataset.BOOKCODE,
        r['code'], '', r.get('card', ''), '',
        r['subject'], plain(r['chapter']), plain(r['big']),
        plain(r['mid']), plain(r['small']),
        r['freq'], r['level'], r['kind'], '', note(r),
        rich(r['t']),
        '<br>'.join(rich(x, False) for x in r['qb']),
        ('%s_stem1.png' % r['code']) if r['qfig'] else '',
        rich(r['c'][0], False), '', rich(r['c'][1], False), '',
        rich(r['c'][2], False), '', rich(r['c'][3], False), '',
        r['a'], solution_html(r), wrongs(r['w'], r['a']),
        point(r['kn'], r['key'], r['mem'], r['old'])
        + lawcite.block(r, details=False),
    ]


def write_xlsx(path, recs, sheet_from_template=True):
    if sheet_from_template:
        wb = openpyxl.load_workbook(TEMPLATE)
        ws = wb['문항등록']
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '문항등록'
        header(ws)
    top = Alignment(vertical='top', wrap_text=True)
    for i, r in enumerate(recs):
        for j, v in enumerate(row_values(r)):
            c = ws.cell(row=i + 3, column=j + 1, value=v)
            c.alignment = top
        ws.row_dimensions[i + 3].height = 160
    wb.save(path)


# ------------------------------------------------------------------ 검수본

CSS = """:root{--ink:#1f2937;--mut:#64748b;--line:#e2e8f0;--accent:#ea580c;--accbg:#fff7ed}
*{box-sizing:border-box}
body{font-family:'Pretendard','Malgun Gothic',system-ui,sans-serif;color:var(--ink);max-width:860px;margin:0 auto;padding:24px 18px 80px;line-height:1.7}
h1{font-size:20px;background:linear-gradient(90deg,#ea580c,#f59e0b);color:#fff;padding:14px 18px;border-radius:12px}
.meta{color:var(--mut);font-size:13px;margin:6px 2px 22px}
.subj{font-weight:800;color:var(--accent);border-left:5px solid var(--accent);padding:6px 12px;background:var(--accbg);border-radius:6px;margin:30px 0 14px;font-size:16px}
.q{border:1px solid var(--line);border-radius:12px;padding:16px 18px;margin:14px 0;box-shadow:0 1px 3px rgba(0,0,0,.03)}
.qh{display:flex;align-items:center;gap:10px;margin-bottom:8px;flex-wrap:wrap}
.qno{font-weight:800;color:var(--accent);font-size:18px}
.freq{color:#f59e0b;font-size:13px}
.meta-tag{display:inline-block;color:#fff;font-size:11px;font-weight:700;border-radius:20px;padding:2px 9px;margin-left:2px}
.code{margin-left:auto;color:#94a3b8;font-size:11px;font-family:monospace}
.stem{font-weight:600;margin:4px 0 12px}
.opts{margin:4px 0 8px}
.opt{padding:3px 0}
.opt.ans{color:#059669;font-weight:700}
.lbl{display:inline-block;font-size:12px;font-weight:700;color:#fff;background:var(--accent);border-radius:20px;padding:2px 10px;margin:14px 0 6px}
.box{border:1px solid #cbd5e1;background:#f1f5f9;border-radius:8px;padding:10px 12px;margin:8px 0}
.small{color:var(--mut);font-size:12px}
.topic{color:#94a3b8;font-size:12px;margin-top:8px}
details.law{border:1px solid #e2e8f0;border-radius:8px;margin:10px 0;background:#fcfcfd}
details.law>summary{cursor:pointer;padding:8px 12px;font-size:13px;font-weight:700;color:#475569;list-style:none;display:flex;align-items:center;gap:8px;flex-wrap:wrap}
details.law>summary::-webkit-details-marker{display:none}
details.law[open]>summary{border-bottom:1px solid #eef2f6}
.eff{margin-left:auto;color:#a3aab5;font-size:11px;font-weight:600;letter-spacing:-.2px}
.lawin{padding:4px 14px 12px}
.lawart{margin:9px 0;font-size:13.5px;line-height:1.72}
.lawart .h{font-weight:700;display:flex;gap:8px;align-items:baseline;flex-wrap:wrap}
.lawart .h a{color:#334155;text-decoration:none}
.lawart .h a:hover{text-decoration:underline}
.lawart .t{white-space:normal;color:#334155;margin-top:3px}
.ln{display:block;white-space:pre-wrap;padding:1px 0}
.ln.hit{background:#fffbeb;border-left:3px solid #f59e0b;padding:4px 8px;margin:3px 0 3px -11px;border-radius:0 5px 5px 0}
.here{display:inline-block;margin-left:8px;font-size:10.5px;font-weight:800;color:#b45309;background:#fef3c7;border:1px solid #fde68a;border-radius:20px;padding:0 8px;vertical-align:middle;white-space:nowrap}
.lawart mark{background:#fde68a;color:inherit;padding:0 2px;border-radius:3px;box-decoration-break:clone;-webkit-box-decoration-break:clone}
.lawmiss{font-size:12px;color:#94a3b8;padding:2px 0}
.lawmiss a{color:#94a3b8}
.cardlink{display:inline-block;margin:10px 0 0;font-size:13px;text-decoration:none;background:#ecfeff;border:1px solid #a5f3fc;color:#0e7490;border-radius:20px;padding:4px 12px;font-weight:700}
.warn{margin-top:10px;padding:8px 10px;background:#fef2f2;border:1px solid #fecaca;border-radius:8px;font-size:12px;color:#b91c1c}
.q img{max-width:88%;display:block;margin:10px 0;padding:6px;border:1px solid #eee;border-radius:6px;background:#fff}
.katex{font-size:1.02em}
.katex-display{overflow-x:auto;overflow-y:hidden;padding:4px 0;margin:8px 0}
.q > div{line-height:1.85}
.katex .text{font-family:'Malgun Gothic','Apple SD Gothic Neo','Noto Sans KR','Pretendard',sans-serif}"""

LVCOLOR = {1: '#10b981', 2: '#f59e0b', 3: '#ef4444'}
LVNAME = {1: '쉬움', 2: '보통', 3: '어려움'}
MARK = ['①', '②', '③', '④']


def review_html(recs, year, rno):
    h = ['<!doctype html><html lang=ko><head><meta charset=utf-8>'
         '<meta name=viewport content="width=device-width,initial-scale=1">'
         '<title>%d년 %d회 산업안전기사 CBT 기출 · 검수본</title>' % (year, rno),
         '<link rel=stylesheet href="katex/katex.min.css">',
         '<script defer src="katex/katex.min.js"></script>',
         '<script defer src="katex/auto-render.min.js" onload="renderMathInElement('
         "document.body,{delimiters:[{left:'$$',right:'$$',display:true},"
         "{left:'$',right:'$',display:false}],throwOnError:false})\"></script>",
         '<style>%s</style></head><body>' % CSS,
         '<h1>%d년 %d회 · 산업안전기사 필기 CBT <span style="font-weight:500;'
         'font-size:14px">기출 재배열 전사 검수본</span></h1>' % (year, rno),
         '<div class=meta>지면(PDF)과 대조 검수용 · 수식은 KaTeX 렌더 · 총 '
         '<b id=cnt></b>문항</div>']
    cur = 0
    for r in recs:
        if r['s'] != cur:
            cur = r['s']
            h.append('<div class=subj>제%d과목 %s</div>' % (cur, r['subject']))
        h.append('<div class=q>')
        h.append('<div class=qh><span class=qno>%03d</span>'
                 '<span class=freq>빈출도 %s</span>'
                 '<span class=meta-tag style="background:%s">난이도 %d·%s</span>'
                 '<span class=meta-tag style="background:#6366f1">%s</span>'
                 '<span class=code>%s</span></div>'
                 % (r['n'], '★' * r['freq'] + '☆' * (3 - r['freq']),
                    LVCOLOR[r['level']], r['level'], LVNAME[r['level']],
                    r['kind'], r['code']))
        h.append('<div class=stem>%s</div>' % rich(r['t']))
        if r['qb']:
            h.append('<div class=box>%s</div>'
                     % '<br>'.join(rich(x, False) for x in r['qb']))
        if r['qfig']:
            h.append('<img src="img/%s_stem1.png">' % r['code'])
        h.append('<div class=opts>')
        for i in range(4):
            h.append('<div class="opt%s">%s %s</div>'
                     % (' ans' if i + 1 == r['a'] else '', MARK[i],
                        rich(r['c'][i], False)))
        h.append('</div>')
        h.append('<div class=small>정답: <b style="color:#059669">%d</b></div>' % r['a'])
        h.append('<div class=lbl>해설</div><div>%s</div>' % rich(r['sol']))
        if r['tbl']:
            h.append('<div class=box>%s</div>' % table(r['tbl']))
        if r['fig']:
            h.append('<img src="img/%s_exp1.png">' % r['code'])
            if r['figcap']:
                h.append('<div class=small>%s</div>' % rich(r['figcap'], False))
        wr = wrongs(r['w'], r['a'])
        if wr:
            h.append('<div class=lbl>오답 분석</div><div>%s</div>' % wr)
        pt = point(r['kn'], r['key'], r['mem'], r['old'])
        if pt:
            h.append('<div class=lbl>관련 개념 · 암기</div><div class=box>%s</div>' % pt)
        lw = lawcite.block(r)
        if lw:
            h.append(lw)
        if r.get('card'):
            h.append('<a class=cardlink href="../_암기카드.html#%s" target="_blank">'
                     '🃏 관련 암기카드</a>' % r['card'])
        fl = flagtext(r)
        if fl:
            h.append('<div class=warn>⚠ %s</div>' % esc(' | '.join(fl)))
        h.append('<div class=topic>분류: %s &gt; %s &gt; %s &nbsp;·&nbsp; 원본 %s</div>'
                 % (esc(plain(r['chapter']) or '—'), esc(plain(r['big']) or '—'),
                    esc(plain(r['small'])), esc(r['src'])))
        h.append('</div>')
    h.append('<script>document.getElementById("cnt").textContent="%d"</script>'
             '</body></html>' % len(recs))
    return '\n'.join(h)


INDEX_CSS = """:root{--ink:#1f2937;--mut:#64748b;--line:#e6e9ef;--accent:#ea580c;--accbg:#fff7ed;--card:#fff;--bg:#f8fafc}
@media(prefers-color-scheme:dark){:root{--ink:#e5e7eb;--mut:#9ca3af;--line:#2a2f3a;--card:#161a22;--bg:#0e1117;--accbg:#2a1a10}}
*{box-sizing:border-box}
body{font-family:Pretendard,"Malgun Gothic",system-ui,sans-serif;color:var(--ink);background:var(--bg);max-width:1000px;margin:0 auto;padding:28px 18px 90px;line-height:1.6}
h1{font-size:22px;background:linear-gradient(90deg,#ea580c,#f59e0b);color:#fff;padding:16px 20px;border-radius:14px;margin:0 0 6px}
.sub{color:var(--mut);font-size:13px;margin:0 2px 26px}
.yr{font-weight:800;color:var(--accent);font-size:15px;border-left:5px solid var(--accent);background:var(--accbg);padding:6px 12px;border-radius:6px;margin:26px 0 12px}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(150px,1fr));gap:12px}
a.card{display:block;text-decoration:none;color:inherit;background:var(--card);border:1px solid var(--line);border-radius:12px;padding:14px 16px;box-shadow:0 1px 3px rgba(0,0,0,.04);transition:.15s}
a.card:hover{border-color:var(--accent);transform:translateY(-2px);box-shadow:0 4px 12px rgba(234,88,12,.15)}
.rnd{font-weight:800;font-size:17px}
.cnt{color:var(--mut);font-size:12px;margin-top:6px}
.badge{display:inline-block;font-size:11px;color:#fff;background:var(--accent);border-radius:20px;padding:1px 9px;margin-top:8px}
.stat{display:flex;gap:22px;margin:0 2px 8px;font-size:13px;color:var(--mut)}
.stat b{color:var(--ink);font-size:16px}"""


def index_html(byround):
    tot = sum(len(v) for v in byround.values())
    figs = sum(bool(r['fig']) + bool(r['qfig'])
               for v in byround.values() for r in v)
    h = ['<!doctype html><html lang=ko><head><meta charset=utf-8>',
         '<meta name=viewport content="width=device-width,initial-scale=1">',
         '<title>산업안전기사 필기 CBT 기출 재배열 · 검수본 인덱스</title>',
         '<style>%s</style></head><body>' % INDEX_CSS,
         '<h1>산업안전기사 필기 CBT 기출 재배열 · 전사 검수본</h1>',
         '<div class=sub>2011~2022년 실제 기출을 출제 비중대로 다시 엮은 24개 회차 · '
         '각 120문항 · 6과목 · 지면(PDF) 대조 검수용</div>',
         '<div class=stat><span><b>24</b> 회차</span><span><b>%s</b> 문항</span>'
         '<span><b>%d</b> 그림</span></div>' % ('{:,}'.format(tot), figs)]
    for y in dataset.YEARS:
        h.append('<div class=yr>%d년</div><div class=grid>' % y)
        for r in sorted(byround):
            yy, rn = dataset.round_label(r)
            if yy != y:
                continue
            recs = byround[r]
            nf = sum(bool(x['fig']) + bool(x['qfig']) for x in recs)
            h.append('<a class=card href="%s/%d_%d회_검수본.html">'
                     '<div class=rnd>%d회</div><div class=cnt>120문항</div>'
                     '<div class=badge>그림 %d</div></a>'
                     % (dataset.folder(r), yy, rn, rn, nf))
        h.append('</div>')
    h.append('</body></html>')
    return '\n'.join(h)


# ------------------------------------------------------------------ 암기카드 엑셀

CARDHEAD = ['과목ID', '챕터', '대유형', '중유형', '내용', '코드', '카드ID', '빈출도',
            '난이도', '카드유형', '앞면', '앞면그림', '뒷면', '뒷면그림', '예제문항코드']
CARDGRP = [('분류', 1, 5, '0891B2'), ('코드', 6, 7, '4F46E5'), ('속성', 8, 10, 'E11D48'),
           ('앞면', 11, 12, 'DB2777'), ('뒷면', 13, 14, 'EA580C'),
           ('예제', 15, 15, '059669')]
CARDFILL = ['CFFAFE'] * 5 + ['E0E7FF'] * 2 + ['FFE4E6'] * 3 + \
           ['FCE7F3', 'FEF3C7', 'FFEDD5', 'FEF3C7', 'D1FAE5']
CARDWIDTH = [16, 16, 18, 24, 22, 12, 20, 8, 8, 10, 40, 16, 60, 16, 34]


def write_cards_xlsx(path, cards):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '카드등록'
    for name, c1, c2, rgb in CARDGRP:
        cell = ws.cell(row=1, column=c1, value=name)
        cell.fill = PatternFill('solid', fgColor=rgb)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if c2 > c1:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
    for i, name in enumerate(CARDHEAD):
        c = ws.cell(row=2, column=i + 1, value=name)
        c.fill = PatternFill('solid', fgColor=CARDFILL[i])
        c.font = Font(bold=True, color='1F2937')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i + 1)].width = CARDWIDTH[i]
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = 'G3'
    top = Alignment(vertical='top', wrap_text=True)
    for i, c in enumerate(cards):
        vals = [c['subject'], c['chapter'], c['big'], c['mid'], c['small'],
                c['code'], c['id'], c['freq'], c['level'], c['kind'],
                c['front'], '', c['back'], '', ', '.join(c['codes'])]
        for j, v in enumerate(vals):
            cell = ws.cell(row=i + 3, column=j + 1, value=v)
            cell.alignment = top
        ws.row_dimensions[i + 3].height = 120
    wb.save(path)


# ------------------------------------------------------------------ 첫 화면

MAIN_CSS = INDEX_CSS + """
.hero{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:12px;margin:18px 0 6px}
a.hcard{display:block;text-decoration:none;color:inherit;border-radius:14px;padding:18px 20px;
border:1px solid var(--line);background:var(--card);transition:.15s}
a.hcard:hover{transform:translateY(-2px);box-shadow:0 6px 16px rgba(234,88,12,.16);border-color:var(--accent)}
a.hcard .t{font-weight:800;font-size:17px;margin-bottom:5px}
a.hcard .d{color:var(--mut);font-size:12.5px;line-height:1.5}
a.hcard .ico{font-size:22px;margin-bottom:8px}
.rcard{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:12px 14px}
.rcard .rnd{font-weight:800;font-size:17px}
.rcard .cnt{color:var(--mut);font-size:11.5px;margin:3px 0 9px}
.two{display:grid;grid-template-columns:1fr 1fr;gap:6px}
.two a{display:block;text-align:center;text-decoration:none;border-radius:8px;padding:7px 0;
font-size:12.5px;font-weight:800;border:1px solid transparent}
a.study{background:#ecfeff;border-color:#a5f3fc;color:#0e7490}
a.study:hover{background:#cffafe}
a.exam{background:#fff7ed;border-color:#fed7aa;color:#c2410c}
a.exam:hover{background:#ffedd5}
"""


def main_index(byround, ncards):
    tot = sum(len(v) for v in byround.values())
    h = ['<!doctype html><html lang=ko><head><meta charset=utf-8>',
         '<meta name=viewport content="width=device-width,initial-scale=1">',
         '<title>산업안전기사 필기 CBT</title>',
         '<style>%s</style></head><body>' % MAIN_CSS,
         '<h1>산업안전기사 필기 CBT</h1>',
         '<div class=sub>2011~2022년 실제 기출을 출제 비중대로 다시 엮은 24개 회차 · '
         '각 120문항 · 6과목 &nbsp;·&nbsp; '
         '<b style="color:#0e7490">📖 학습</b> 해설·법령·암기카드를 열어 가며 풀기 &nbsp;/&nbsp; '
         '<b style="color:#c2410c">⏱ 시험</b> 3시간 재고 제출 뒤 채점</div>',
         '<div class=stat><span><b>24</b> 회차</span>'
         '<span><b>%s</b> 문항</span><span><b>%s</b> 암기카드</span></div>'
         % ('{:,}'.format(tot), '{:,}'.format(ncards)),
         '<div class=hero>',
         '<a class=hcard href="_암기카드.html"><div class=ico>🃏</div>'
         '<div class=t>암기카드 %s장</div><div class=d>기출에서 뽑은 개념 카드. '
         '뒤집어 확인하고, 그 개념이 나온 문항으로 바로 건너뜁니다</div></a>'
         % '{:,}'.format(ncards),
         '<a class=hcard href="_index_검수본.html"><div class=ico>📄</div>'
         '<div class=t>전사 검수본</div><div class=d>한 회차를 한 쪽에 펼쳐 놓은 대조용 화면. '
         '문항·해설·그림·수식을 지면과 맞춰 봅니다</div></a>',
         '</div>']
    for y in dataset.YEARS:
        h.append('<div class=yr>%d년</div><div class=grid>' % y)
        for r in sorted(byround):
            yy, rn = dataset.round_label(r)
            if yy != y:
                continue
            h.append('<div class=rcard><div class=rnd>%d회</div>'
                     '<div class=cnt>120문항 · 6과목</div><div class=two>'
                     '<a class=study href="%s/%d_%d회_학습.html">📖 학습</a>'
                     '<a class=exam href="%s/%d_%d회_CBT.html">⏱ 시험</a>'
                     '</div></div>'
                     % (rn, dataset.folder(r), yy, rn, dataset.folder(r), yy, rn))
        h.append('</div>')
    h.append('</body></html>')
    return '\n'.join(h)


# ------------------------------------------------------------------ 실행

def cbt_parts(r):
    """CBT 화면에 심을 문항 조각."""
    return {
        't': rich(r['t']),
        'c': [rich(x, False) for x in r['c']],
        'cond': '<br>'.join(rich(x, False) for x in r['qb']),
        'img': ('%s_stem1.png' % r['code']) if r['qfig'] else '',
        'sol': solution_html(r).replace('src="%s_exp1.png"' % r['code'],
                                        'src="img/%s_exp1.png"' % r['code']),
        'w': wrongs(r['w'], r['a']),
        'pt': point(r['kn'], r['key'], r['mem'], r['old']),
        'law': lawcite.block(r),
        'card': r.get('card', ''),
    }


def main():
    recs = dataset.build()
    cards, bycode = cardmod.build(recs)
    cardmap = {c['id']: c for c in cards}
    for r in recs:
        r['card'] = bycode.get(r['code'], '')
    print('암기카드 %d장 · 문항 연결 %d' % (len(cards), sum(1 for r in recs if r['card'])))

    byround = collections.defaultdict(list)
    for r in recs:
        byround[r['round']].append(r)

    make_template()
    for rd in sorted(byround):
        rs = byround[rd]
        y, rn = dataset.round_label(rd)
        d = os.path.join(OUT, dataset.folder(rd))
        img = os.path.join(d, 'img')
        os.makedirs(img, exist_ok=True)
        for f in os.listdir(img):
            os.remove(os.path.join(img, f))
        kx = os.path.join(d, 'katex')
        if not os.path.isdir(kx):
            shutil.copytree(KATEX_SRC, kx)
        for r in rs:
            if r['qfig']:
                shutil.copyfile(os.path.join(dataset.FIG, r['qfig']),
                                os.path.join(img, '%s_stem1.png' % r['code']))
            if r['fig']:
                shutil.copyfile(os.path.join(dataset.FIG, r['fig']),
                                os.path.join(img, '%s_exp1.png' % r['code']))
        write_xlsx(os.path.join(d, '산업안전기사_%d_%d회_문항등록.xlsx' % (y, rn)), rs)
        with open(os.path.join(d, '%d_%d회_검수본.html' % (y, rn)), 'w',
                  encoding='utf-8') as fp:
            fp.write(review_html(rs, y, rn))
        for mode, suffix in (('exam', 'CBT'), ('study', '학습')):
            with open(os.path.join(d, '%d_%d회_%s.html' % (y, rn, suffix)), 'w',
                      encoding='utf-8') as fp:
                fp.write(cbt_app.build_round(rs, y, rn, rd, cbt_parts,
                                             cardmap, mode))
        print('  %s  %d문항' % (dataset.folder(rd), len(rs)))

    write_xlsx(os.path.join(OUT, '산업안전기사_CBT_24회차_통합.xlsx'), recs)

    kx = os.path.join(OUT, 'katex')
    if not os.path.isdir(kx):
        shutil.copytree(KATEX_SRC, kx)
    write_cards_xlsx(os.path.join(
        OUT, '산업안전기사_암기카드_%d장.xlsx' % len(cards)), cards)
    with open(os.path.join(OUT, '_암기카드.html'), 'w', encoding='utf-8') as fp:
        fp.write(card_view.build(cards, dataset.SUBJ))
    with open(os.path.join(OUT, '_index_검수본.html'), 'w', encoding='utf-8') as fp:
        fp.write(index_html(byround))
    with open(os.path.join(OUT, '_index.html'), 'w', encoding='utf-8') as fp:
        fp.write(main_index(byround, len(cards)))
    print('완료 — 회차 %d · 문항 %d · 카드 %d' % (len(byround), len(recs), len(cards)))


if __name__ == '__main__':
    main()
